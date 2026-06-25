import openpyxl
from django.contrib import admin
from django.db import models
from django.http import HttpResponse
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
from jalali_date import date2jalali, datetime2jalali
from jalali_date.fields import JalaliDateField, SplitJalaliDateTimeField
from jalali_date.widgets import AdminJalaliDateWidget, AdminSplitJalaliDateTime
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import Category, InPersonClass, InPersonClassRegistration, TimeRange

_JALALI_FORMFIELD_OVERRIDES = {
    models.DateField: {'form_class': JalaliDateField, 'widget': AdminJalaliDateWidget},
    models.DateTimeField: {'form_class': SplitJalaliDateTimeField, 'widget': AdminSplitJalaliDateTime},
}


def export_registrations_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ثبت‌نام‌ها"
    ws.sheet_view.rightToLeft = True

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin')
    full_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')

    def write_cell(row, col, value, border=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = center
        if border:
            cell.border = border
        return cell

    qs = queryset.prefetch_related(
        'bought_by', 'bought_by__profile',
        'in_person_class', 'time_range',
    )

    current_row = 1
    user_headers = [
        "نام کاربری", "نام کامل", "شماره تلفن", "ایمیل",
        "کد ملی", "جنسیت", "شهر", "تلگرام", "ایتا", "اینستاگرام",
    ]

    for reg in qs:
        start_jalali = date2jalali(reg.start_date).strftime('%Y/%m/%d')
        end_jalali = date2jalali(reg.end_date).strftime('%Y/%m/%d')
        price_str = f"{reg.price:,} تومان"
        new_price_str = f"{reg.new_price:,} تومان" if reg.new_price else "—"

        meta_rows = [
            ("نام کلاس", reg.in_person_class.title),
            ("بازه زمانی", reg.time_range.label),
            ("تاریخ شروع", start_jalali),
            ("تاریخ پایان", end_jalali),
            ("قیمت", price_str),
            ("قیمت جدید", new_price_str),
        ]
        for label, value in meta_rows:
            label_cell = write_cell(current_row, 1, label)
            label_cell.font = Font(bold=True)
            write_cell(current_row, 2, value)
            current_row += 1

        current_row += 1

        for col_idx, header in enumerate(user_headers, start=1):
            cell = write_cell(current_row, col_idx, header, border=full_border)
            cell.font = header_font
            cell.fill = header_fill
        current_row += 1

        for user in reg.bought_by.all():
            profile = getattr(user, 'profile', None)
            gender_display = profile.get_gender_display() if profile and profile.gender else ""
            row_data = [
                user.username,
                user.name,
                user.phone,
                user.email or "",
                getattr(profile, 'national_code', "") or "",
                gender_display,
                getattr(profile, 'city', "") or "",
                getattr(profile, 'telegram_id', "") or "",
                getattr(profile, 'eitaa_id', "") or "",
                getattr(profile, 'instagram_id', "") or "",
            ]
            for col_idx, value in enumerate(row_data, start=1):
                write_cell(current_row, col_idx, value, border=full_border)
            current_row += 1

        current_row += 2

    col_widths = [18, 20, 15, 25, 14, 10, 14, 16, 16, 18]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="registrations.xlsx"'
    wb.save(response)
    return response


export_registrations_excel.short_description = "خروجی اکسل از ثبت‌نام‌های انتخاب‌شده"


@admin.register(TimeRange)
class TimeRangeAdmin(admin.ModelAdmin):
    list_display = ('label',)
    search_fields = ('label',)
    ordering = ('label',)


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ('name', 'slug')
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ('name', 'slug')
    ordering = ('name',)


@admin.register(InPersonClass)
class InPersonClassAdmin(admin.ModelAdmin):
    formfield_overrides = _JALALI_FORMFIELD_OVERRIDES
    list_display = (
        'title',
        'start_date_jalali',
        'end_date_jalali',
        'price_display',
        'discount_display',
        'is_visible',
        'created_at_jalali',
    )

    list_filter = ('is_visible', 'start_date', 'categories')
    search_fields = ('title',)
    ordering = ('-start_date',)
    date_hierarchy = 'start_date'
    list_per_page = 50

    readonly_fields = ('id', 'created_at', 'updated_at')
    filter_horizontal = ('available_times', 'categories')

    fieldsets = (
        (
            _('Core Information'),
            {
                'fields': ('id', 'title', 'thumbnail', 'is_visible'),
            },
        ),
        (
            _('Description'),
            {
                'fields': ('short_description',),
            },
        ),
        (
            _('Pricing'),
            {
                'fields': ('price', 'new_price'),
            },
        ),
        (
            _('Schedule'),
            {
                'fields': ('start_date', 'end_date', 'available_times'),
            },
        ),
        (
            _('Online Session'),
            {
                'fields': ('meeting_url',),
                'description': _('Set this link if the class will be held online (e.g. due to weather or other issues).'),
            },
        ),
        (
            _('Categories'),
            {
                'fields': ('categories',),
            },
        ),
        (
            _('Timestamps'),
            {
                'fields': ('created_at', 'updated_at'),
            },
        ),
    )

    @admin.display(description=_('Start Date'))
    def start_date_jalali(self, obj):
        return date2jalali(obj.start_date).strftime('%Y/%m/%d')

    @admin.display(description=_('End Date'))
    def end_date_jalali(self, obj):
        return date2jalali(obj.end_date).strftime('%Y/%m/%d')

    @admin.display(description=_('Created At'))
    def created_at_jalali(self, obj):
        return datetime2jalali(timezone.localtime(obj.created_at)).strftime('%Y/%m/%d %H:%M')

    @admin.display(description=_('Price'))
    def price_display(self, obj):
        effective = obj.new_price if obj.new_price else obj.price
        return f'{effective:,} تومان'

    @admin.display(description=_('Discount'))
    def discount_display(self, obj):
        d = obj.discount
        if d:
            return format_html(
                '<span style="padding:3px 8px; border-radius:10px; '
                'background:#dc3545; color:#fff; font-weight:600;">{}%</span>',
                d,
            )
        return '—'


@admin.register(InPersonClassRegistration)
class InPersonClassRegistrationAdmin(admin.ModelAdmin):
    formfield_overrides = _JALALI_FORMFIELD_OVERRIDES
    list_display = (
        'in_person_class',
        'time_range',
        'start_date_jalali',
        'end_date_jalali',
        'price',
        'registered_count',
        'created_at_jalali',
    )

    actions = [export_registrations_excel]

    list_filter = ('in_person_class',)
    search_fields = ('in_person_class__title', 'time_range__label')
    list_select_related = ('in_person_class', 'time_range')
    ordering = ('-created_at',)
    list_per_page = 50

    readonly_fields = ('id', 'start_date', 'end_date', 'price', 'new_price', 'created_at')
    filter_horizontal = ('bought_by',)

    @admin.display(description=_('Start Date'))
    def start_date_jalali(self, obj):
        return date2jalali(obj.start_date).strftime('%Y/%m/%d')

    @admin.display(description=_('End Date'))
    def end_date_jalali(self, obj):
        return date2jalali(obj.end_date).strftime('%Y/%m/%d')

    @admin.display(description=_('Created At'))
    def created_at_jalali(self, obj):
        return datetime2jalali(timezone.localtime(obj.created_at)).strftime('%Y/%m/%d %H:%M')

    @admin.display(description=_('Registered'))
    def registered_count(self, obj):
        return obj.bought_by.count()

    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related('bought_by')

    fieldsets = (
        (
            _('Registration'),
            {
                'fields': ('id', 'in_person_class', 'time_range'),
            },
        ),
        (
            _('Snapshot at Registration'),
            {
                'fields': ('start_date', 'end_date', 'price', 'new_price'),
            },
        ),
        (
            _('Students'),
            {
                'fields': ('bought_by',),
            },
        ),
        (
            _('Timestamps'),
            {
                'fields': ('created_at',),
            },
        ),
    )
